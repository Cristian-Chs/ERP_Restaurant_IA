"""
Vistas de Analytics con Pandas para Automatización del Restaurante
Incluye: Control de Inventario Dinámico, Matriz BCG, Proyección de Compras, Exportes Excel
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from django.http import HttpResponse
from django.db.models import Sum, Count, F, Q
from decimal import Decimal
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference, LineChart

from .models import Product, Recipe, Ingredient, InventoryMovement, Employee, PayrollPayment
from bot.models import Order


class DynamicCostRecalculationView(APIView):
    """
    Recalcula el costo de todos los productos basándose en sus recetas y costos de ingredientes.
    Detecta inflación y genera alertas.
    """
    permission_classes = [permissions.IsAdminUser]
    
    def post(self, request):
        # Obtener todos los productos con recetas
        products_with_recipes = Product.objects.filter(recetas__isnull=False).distinct()
        
        changes = []
        total_inflation = 0
        
        for product in products_with_recipes:
            # Calcular costo actual basado en recetas
            old_cost = product.cost_price
            new_cost = sum([recipe.get_cost() for recipe in product.recetas.all()])
            
            if new_cost != old_cost:
                inflation_percent = ((new_cost - old_cost) / old_cost * 100) if old_cost > 0 else 0
                
                # Actualizar el producto
                product.cost_price = new_cost
                product.save()
                
                changes.append({
                    'product_id': product.id,
                    'product_name': product.name,
                    'old_cost': float(old_cost),
                    'new_cost': float(new_cost),
                    'difference': float(new_cost - old_cost),
                    'inflation_percent': round(inflation_percent, 2),
                    'current_price': float(product.price),
                    'margin': float(product.price - new_cost),
                    'margin_percent': round(((product.price - new_cost) / product.price * 100) if product.price > 0 else 0, 2)
                })
                
                total_inflation += inflation_percent
        
        # Calcular inflación promedio
        avg_inflation = (total_inflation / len(changes)) if changes else 0
        
        # Generar alertas
        alerts = []
        if avg_inflation > 5:
            alerts.append({
                'type': 'warning',
                'message': f' Tus costos subieron un {avg_inflation:.1f}% en promedio. Considera ajustar precios.'
            })
        
        # Productos con margen bajo
        low_margin_products = [c for c in changes if c['margin_percent'] < 20]
        if low_margin_products:
            alerts.append({
                'type': 'danger',
                'message': f' {len(low_margin_products)} productos tienen margen menor al 20%. Revisa precios urgentemente.'
            })
        
        return Response({
            'success': True,
            'products_updated': len(changes),
            'average_inflation': round(avg_inflation, 2),
            'changes': changes,
            'alerts': alerts
        })


class BCGMatrixView(APIView):
    """
    Clasifica productos en la Matriz BCG (Boston Consulting Group):
    - Estrellas: Alta venta + Alto margen
    - Caballos: Alta venta + Bajo margen
    - Incógnitas: Baja venta + Alto margen
    - Perros: Baja venta + Bajo margen
    """
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request):
        # Obtener ventas de los últimos 30 días
        thirty_days_ago = datetime.now() - timedelta(days=30)
        
        # Obtener datos de ventas por producto
        sales_data = Order.objects.filter(
            payment_status='payment_approved',
            fecha__gte=thirty_days_ago
        ).values('item').annotate(
            total_sales=Count('id'),
            total_revenue=Sum('precio')
        )
        
        # Crear DataFrame con Pandas
        df_sales = pd.DataFrame(list(sales_data))
        
        if df_sales.empty:
            return Response({
                'message': 'No hay suficientes datos de ventas para generar la matriz BCG',
                'categories': {'stars': [], 'horses': [], 'questions': [], 'dogs': []}
            })
        
        # Obtener información de productos
        products = Product.objects.all().values('id', 'name', 'price', 'cost_price', 'category')
        df_products = pd.DataFrame(list(products))
        
        # Merge de datos
        df = pd.merge(
            df_sales,
            df_products,
            left_on='item',
            right_on='name',
            how='left'
        )
        
        # Calcular margen
        df['margin'] = df['price'] - df['cost_price']
        df['margin_percent'] = (df['margin'] / df['price'] * 100).fillna(0)
        
        # Calcular percentiles para clasificación
        sales_median = df['total_sales'].median()
        margin_median = df['margin_percent'].median()
        
        # Clasificar productos
        def classify_product(row):
            high_sales = row['total_sales'] >= sales_median
            high_margin = row['margin_percent'] >= margin_median
            
            if high_sales and high_margin:
                return 'star'
            elif high_sales and not high_margin:
                return 'horse'
            elif not high_sales and high_margin:
                return 'question'
            else:
                return 'dog'
        
        df['category_bcg'] = df.apply(classify_product, axis=1)
        
        # Preparar respuesta
        result = {
            'stars': df[df['category_bcg'] == 'star'].to_dict('records'),
            'horses': df[df['category_bcg'] == 'horse'].to_dict('records'),
            'questions': df[df['category_bcg'] == 'question'].to_dict('records'),
            'dogs': df[df['category_bcg'] == 'dog'].to_dict('records')
        }
        
        # Convertir Decimals a float para JSON
        for category in result:
            for item in result[category]:
                for key, value in item.items():
                    if isinstance(value, Decimal):
                        item[key] = float(value)
        
        # Agregar recomendaciones
        recommendations = []
        if len(result['stars']) > 0:
            recommendations.append({
                'type': 'success',
                'message': f' Tienes {len(result["stars"])} productos estrella. ¡Mantenlos en el menú y promociónalos!'
            })
        
        if len(result['horses']) > 0:
            recommendations.append({
                'type': 'warning',
                'message': f' {len(result["horses"])} productos se venden mucho pero dan poco margen. Considera subir ligeramente el precio.'
            })
        
        if len(result['dogs']) > 0:
            recommendations.append({
                'type': 'danger',
                'message': f' {len(result["dogs"])} productos no se venden y dan poco margen. Considera eliminarlos del menú.'
            })
        
        return Response({
            'categories': result,
            'recommendations': recommendations,
            'thresholds': {
                'sales_median': float(sales_median),
                'margin_median': float(margin_median)
            }
        })


class PurchasePredictionView(APIView):
    """
    Analiza el histórico de ventas y predice las necesidades de compra de ingredientes
    para los próximos 7 días.
    """
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request):
        # Obtener ventas de los últimos 60 días
        sixty_days_ago = datetime.now() - timedelta(days=60)
        
        orders = Order.objects.filter(
            payment_status='payment_approved',
            fecha__gte=sixty_days_ago
        ).values('item', 'fecha')
        
        df_orders = pd.DataFrame(list(orders))
        
        if df_orders.empty:
            return Response({
                'message': 'No hay suficientes datos históricos para hacer predicciones',
                'suggestions': []
            })
        
        # Convertir fecha a datetime
        df_orders['fecha'] = pd.to_datetime(df_orders['fecha'])
        df_orders['day_of_week'] = df_orders['fecha'].dt.dayofweek
        df_orders['week'] = df_orders['fecha'].dt.isocalendar().week
        
        # Contar ventas por producto y día de semana
        sales_by_day = df_orders.groupby(['item', 'day_of_week']).size().reset_index(name='count')
        
        # Calcular promedio de ventas por día de semana
        avg_sales = sales_by_day.groupby(['item', 'day_of_week'])['count'].mean().reset_index()
        
        # Obtener recetas
        recipes = Recipe.objects.select_related('product', 'ingredient').all()
        
        suggestions = []
        ingredient_totals = {}
        
        # Predecir para los próximos 7 días
        today = datetime.now()
        for i in range(7):
            future_date = today + timedelta(days=i)
            day_of_week = future_date.weekday()
            
            # Para cada producto, predecir ventas
            for product_name in avg_sales['item'].unique():
                predicted_sales = avg_sales[
                    (avg_sales['item'] == product_name) & 
                    (avg_sales['day_of_week'] == day_of_week)
                ]['count'].values
                
                if len(predicted_sales) > 0:
                    qty = predicted_sales[0]
                    
                    # Buscar recetas del producto
                    product_recipes = recipes.filter(product__name=product_name)
                    
                    for recipe in product_recipes:
                        ingredient_name = recipe.ingredient.nombre
                        needed_qty = float(recipe.quantity) * qty
                        
                        if ingredient_name not in ingredient_totals:
                            ingredient_totals[ingredient_name] = {
                                'quantity': 0,
                                'unit': recipe.ingredient.unit,
                                'cost_per_unit': float(recipe.ingredient.cost)
                            }
                        
                        ingredient_totals[ingredient_name]['quantity'] += needed_qty
        
        # Preparar lista de compras
        for ingredient, data in ingredient_totals.items():
            suggestions.append({
                'ingredient': ingredient,
                'quantity': round(data['quantity'], 2),
                'unit': data['unit'],
                'cost_per_unit': data['cost_per_unit'],
                'total_cost': round(data['quantity'] * data['cost_per_unit'], 2)
            })
        
        # Ordenar por costo total descendente
        suggestions.sort(key=lambda x: x['total_cost'], reverse=True)
        
        total_investment = sum([s['total_cost'] for s in suggestions])
        
        return Response({
            'success': True,
            'period': '7 días',
            'suggestions': suggestions,
            'total_investment': round(total_investment, 2),
            'message': f'Necesitarás invertir aproximadamente ${total_investment:.2f} en ingredientes para los próximos 7 días'
        })


class ExportExcelFinancialView(APIView):
    """
    Exporta un reporte financiero completo en Excel con múltiples hojas,
    fórmulas nativas y formato profesional.
    """
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request):
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # === HOJA 1: Ventas Diarias ===
        ws_sales = wb.create_sheet("Ventas Diarias")
        
        # Obtener ventas de los últimos 30 días
        thirty_days_ago = datetime.now() - timedelta(days=30)
        orders = Order.objects.filter(
            payment_status='payment_approved',
            fecha__gte=thirty_days_ago
        ).values('fecha', 'precio')
        
        df_sales = pd.DataFrame(list(orders))
        
        if not df_sales.empty:
            df_sales['fecha'] = pd.to_datetime(df_sales['fecha']).dt.date
            daily_sales = df_sales.groupby('fecha')['precio'].sum().reset_index()
            
            # Headers
            ws_sales['A1'] = 'Fecha'
            ws_sales['B1'] = 'Ventas ($)'
            ws_sales['A1'].font = Font(bold=True)
            ws_sales['B1'].font = Font(bold=True)
            
            # Datos
            for idx, row in daily_sales.iterrows():
                ws_sales[f'A{idx+2}'] = row['fecha']
                ws_sales[f'B{idx+2}'] = float(row['precio'])
            
            # Totales con fórmula
            last_row = len(daily_sales) + 2
            ws_sales[f'A{last_row}'] = 'TOTAL'
            ws_sales[f'A{last_row}'].font = Font(bold=True)
            ws_sales[f'B{last_row}'] = f'=SUM(B2:B{last_row-1})'
            ws_sales[f'B{last_row}'].font = Font(bold=True)
            
            # Promedio
            ws_sales[f'A{last_row+1}'] = 'PROMEDIO'
            ws_sales[f'B{last_row+1}'] = f'=AVERAGE(B2:B{last_row-1})'
        
        # === HOJA 2: Nómina ===
        ws_payroll = wb.create_sheet("Nómina")
        
        payroll = PayrollPayment.objects.select_related('employee').all().order_by('-payment_date')[:50]
        
        ws_payroll['A1'] = 'Empleado'
        ws_payroll['B1'] = 'Rol'
        ws_payroll['C1'] = 'Monto'
        ws_payroll['D1'] = 'Fecha'
        
        for col in ['A1', 'B1', 'C1', 'D1']:
            ws_payroll[col].font = Font(bold=True)
        
        for idx, payment in enumerate(payroll, start=2):
            ws_payroll[f'A{idx}'] = payment.employee.name
            ws_payroll[f'B{idx}'] = payment.employee.get_role_display()
            ws_payroll[f'C{idx}'] = float(payment.amount)
            ws_payroll[f'D{idx}'] = payment.payment_date
        
        if payroll.count() > 0:
            last_row = payroll.count() + 2
            ws_payroll[f'A{last_row}'] = 'TOTAL PAGADO'
            ws_payroll[f'A{last_row}'].font = Font(bold=True)
            ws_payroll[f'C{last_row}'] = f'=SUM(C2:C{last_row-1})'
            ws_payroll[f'C{last_row}'].font = Font(bold=True)
        
        # === HOJA 3: Productos (Matriz BCG) ===
        ws_products = wb.create_sheet("Productos BCG")
        
        products = Product.objects.all()
        
        ws_products['A1'] = 'Producto'
        ws_products['B1'] = 'Precio Venta'
        ws_products['C1'] = 'Costo'
        ws_products['D1'] = 'Margen'
        ws_products['E1'] = 'Margen %'
        
        for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
            ws_products[col].font = Font(bold=True)
        
        for idx, product in enumerate(products, start=2):
            ws_products[f'A{idx}'] = product.name
            ws_products[f'B{idx}'] = float(product.price)
            ws_products[f'C{idx}'] = float(product.cost_price)
            ws_products[f'D{idx}'] = f'=B{idx}-C{idx}'
            ws_products[f'E{idx}'] = f'=D{idx}/B{idx}*100'
        
        # === HOJA 4: Ingredientes ===
        ws_ingredients = wb.create_sheet("Ingredientes")
        
        ingredients = Ingredient.objects.all()
        
        ws_ingredients['A1'] = 'Ingrediente'
        ws_ingredients['B1'] = 'Costo por Unidad'
        ws_ingredients['C1'] = 'Unidad'
        ws_ingredients['D1'] = 'Disponible como Extra'
        
        for col in ['A1', 'B1', 'C1', 'D1']:
            ws_ingredients[col].font = Font(bold=True)
        
        for idx, ingredient in enumerate(ingredients, start=2):
            ws_ingredients[f'A{idx}'] = ingredient.nombre
            ws_ingredients[f'B{idx}'] = float(ingredient.cost)
            ws_ingredients[f'C{idx}'] = ingredient.unit
            ws_ingredients[f'D{idx}'] = 'Sí' if ingredient.disponible_como_extra else 'No'
        
        # Guardar en memoria
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta HTTP
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Reporte_Financiero_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
